import sys

from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font
)
import os
from taiga import TaigaAPI
from datetime import datetime, timedelta
import time
from dotenv import load_dotenv
import argparse
from openpyxl.workbook import Workbook
import getpass
import itertools
import threading

now = datetime.now()


def make_color_lighter(color, percent):
    ar = []
    ar.append(int(min(255, color[0] * (1 + percent / 100))))
    ar.append(int(min(255, color[1] * (1 + percent / 100))))
    ar.append(int(min(255, color[2] * (1 + percent / 100))))
    return ar


def hex_to_rgb(hex):
    return list(int(hex[i:i + 2], 16) for i in (0, 2, 4))


done = False


def generateDocument(user_story, tasks, status_task, status_user_story, monday):
    print('making a document...' + '\033[0m')
    # Alias of a statuses
    status_associations = {'87': 'Нужно больше данных', '86': 'Закрыто', '88': 'Провал', '84': 'В процессе',
                           '83': 'Новое'}

    current_time = now.strftime('%d-%m-%Y')

    wb = Workbook()
    ws = wb.active

    font_name = 'Calibri'

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    alignment = Alignment(
        horizontal='left',
        vertical='bottom',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=True,
        indent=0
    )

    space = 2
    user_story_index = 0
    dynamic_bold = Font(
        name=font_name,
        size=18,
        bold=True,
    )

    little_bold = Font(
        name=font_name,
        size=12,
        bold=True,
    )

    hyper_link_font = Font(
        name=font_name,
        size=12,
        color='0000EE',
        underline='single'
    )

    thin_border_head = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

    tmp_sub = ""

    task = 0
    while task < (len(tasks)):
        bool_in_time_task = datetime.timestamp(datetime.strptime(vars(tasks[task])['modified_date'],
                                                                 '%Y-%m-%dT%H:%M:%S.%fZ')) >= monday

        bool_check_user_story_status = status_user_story in str(
            user_story[vars(tasks[task])['user_story_extra_info']['id']]['status'])

        bool_check_task_status = status_task in str(vars(tasks[task])['status'])

        all_bools = bool_in_time_task and bool_check_user_story_status and bool_check_task_status

        if all_bools:

            tmp_ref = ''

            try:

                # Name of a story
                if tmp_sub != str(vars(tasks[task])['user_story_extra_info']['subject']) and all_bools:
                    ws['A' + str(user_story_index + space)] = str(vars(tasks[task])['user_story_extra_info']['subject'])
                    ws['A' + str(user_story_index + space)].font = dynamic_bold
                    ws['A' + str(user_story_index + space)].alignment = alignment

                    ws.merge_cells(start_row=user_story_index + space, start_column=1, end_row=user_story_index + space,
                                   end_column=2)
                    ws.cell(row=user_story_index + space, column=1).border = thin_border_head
                    ws.cell(row=user_story_index + space, column=2).border = thin_border_head
                    ws.cell(row=user_story_index + space, column=3).border = thin_border_head
                    rd = ws.row_dimensions[user_story_index + space]
                    rd.height = 45

                    actual_color = str(user_story[vars(tasks[task])['user_story_extra_info']['id']]['color']).replace(
                        '#', '')

                    dynamic_color = PatternFill(start_color=actual_color,
                                                end_color=actual_color,
                                                fill_type='solid')

                    ws['C' + str(user_story_index + space)] = str(
                        user_story[vars(tasks[task])['user_story_extra_info']['id']]['name'])
                    tmp_ref = vars(tasks[task])['user_story_extra_info']['ref']

                    ws['C' + str(user_story_index + space)].fill = dynamic_color
                    ws['C' + str(user_story_index + space)].font = dynamic_bold
                    ws['C' + str(user_story_index + space)].alignment = alignment

                    # Names of a header of a story
                    ws['A' + str(user_story_index + space + 1)] = 'Задача'
                    ws['A' + str(user_story_index + space + 1)].font = dynamic_bold
                    ws['A' + str(user_story_index + space + 1)].alignment = alignment
                    ws.cell(row=user_story_index + space + 1, column=1).border = thin_border_head

                    ws['B' + str(user_story_index + space + 1)] = 'Статус'
                    ws['B' + str(user_story_index + space + 1)].font = dynamic_bold
                    ws['B' + str(user_story_index + space + 1)].alignment = alignment
                    ws.cell(row=user_story_index + space + 1, column=2).border = thin_border_head

                    # Test of adding of a date
                    ws['C' + str(user_story_index + space + 1)] = 'Дата'
                    ws['C' + str(user_story_index + space + 1)].font = dynamic_bold
                    ws['C' + str(user_story_index + space + 1)].alignment = alignment
                    ws.cell(row=user_story_index + space + 1, column=3).border = thin_border_head
                    task_index = 0
                    tmp_sub = str(vars(tasks[task])['user_story_extra_info']['ref'])
                # Tasks`s names, statuses and colors
                while tmp_sub == str(vars(tasks[task])['user_story_extra_info']['ref']):

                    if status_task in str(vars(tasks[task])['status']) and datetime.timestamp(
                            datetime.strptime(vars(tasks[task])['modified_date'],
                                              '%Y-%m-%dT%H:%M:%S.%fZ')) > monday:
                        actual_color = hex_to_rgb(str(vars(tasks[task])['status_extra_info']['color']).replace('#', ''))

                        more_light_color = make_color_lighter(actual_color, 0)

                        color = '{:02x}{:02x}{:02x}'.format(*more_light_color)
                        dynamic_color = PatternFill(start_color=color,
                                                    end_color=color,
                                                    fill_type='solid')

                        ws['A' + str(user_story_index + space + 2)] = str(vars(tasks[task])['subject'])
                        ws['A' + str(user_story_index + space + 2)].alignment = alignment
                        ws.cell(row=user_story_index + space + 2, column=1).border = thin_border_head

                        ws['B' + str(user_story_index + space + 2)] = str(
                            vars(tasks[task])['status_extra_info']['name'])
                        ws['B' + str(user_story_index + space + 2)].alignment = alignment
                        ws['B' + str(user_story_index + space + 2)].fill = dynamic_color
                        ws.cell(row=user_story_index + space + 2, column=2).border = thin_border_head

                        # Cell of a date
                        datetime_object = datetime.strptime(str(vars(tasks[task])['modified_date']).split('.')[0],
                                                            '%Y-%m-%dT%H'
                                                            ':%M:%S')
                        four_hours_ago = datetime_object + timedelta(hours=4)

                        ws['C' + str(user_story_index + space + 2)] = four_hours_ago
                        ws['C' + str(user_story_index + space + 2)].alignment = alignment
                        ws.cell(row=user_story_index + space + 2, column=3).border = thin_border_head

                        space = space + 1

                        task_index = task_index + 1
                    task = task + 1

            except:
                print()
                # Bottom of a user_story
            ws['B' + str(user_story_index + space + 2)] = 'Всего: ' + str(task_index)
            ws['B' + str(user_story_index + space + 2)].font = little_bold
            ws['B' + str(user_story_index + space + 2)].alignment = alignment

            ws['A' + str(user_story_index + space + 2)].hyperlink = (
                    'https://taiga.psuti.ru/project/development'
                    '-department/us/' + str(tmp_ref))

            ws['A' + str(user_story_index + space + 2)].font = hyper_link_font
            ws['A' + str(user_story_index + space + 2)].alignment = alignment
            # Adding a results`s border
            ws.cell(row=user_story_index + space + 2, column=1).border = thin_border_head
            ws.cell(row=user_story_index + space + 2, column=2).border = thin_border_head
            ws.cell(row=user_story_index + space + 2, column=3).border = thin_border_head

            ws.merge_cells(start_row=user_story_index + space + 2, start_column=2,
                           end_row=user_story_index + space + 2,
                           end_column=3)
            space = space + 1
            user_story_index = user_story_index + 3

            # Correcting of index
            task = task - 1
        task = task + 1

    wb.save('Отчет-' + current_time + '.xlsx')

def animate():
    for c in itertools.cycle(["⢿", "⣻", "⣽", "⣾", "⣷", "⣯", "⣟", "⡿"]):
        if done:
            break
        sys.stdout.write('\rloading ' + c)
        sys.stdout.flush()
        time.sleep(0.1)
    sys.stdout.write('\r')


if __name__ == '__main__':
    # Adding of args
    parser = argparse.ArgumentParser(description='All parameters',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-t', '--task', type=str, default='',
                        help='fetch by status of a task: 86 - closed; '
                             '88 - fail; 84 - In progress; 87 - Needs Info'
                             ' ; 83 - New; 85 - Ready for test;')

    parser.add_argument('-u', '--userstory', type=str, default='',
                        help='fetch by status of user`s story: 106 - Archived; '
                             '105 - Done; 109 - Waiting for response; 108 - Code'
                             ' review; 103 - In work; 101 - New;')

    parser.add_argument('-l', '--logs', nargs='*', default='1', help='show a logs of url, response and header')
    parser.add_argument('-f', '--fullarr', nargs='*', default='1', help='show all user`s story')
    parser.add_argument('-a', '--all', nargs='*', default='1', help='show all weeks')
    parser.add_argument('-d', '--data', type=str, default='', help='pick custom date, format writing id d-m-y')

    args = parser.parse_args()
    config = vars(args)

    # BEGINNING OF MAIN FUNC
    load_dotenv()

    tasks = ''
    user_story = ''

    # Loading animation start
    t = threading.Thread(target=animate)
    t.start()

    try:
        api = TaigaAPI(host='https://taiga.psuti.ru', token=str(os.getenv('AUTH_TOKEN')))

        print('\033[92m' + '\rGetting info from api...')
        tasks = api.tasks.list(project=17)
    except:
        done = True
        api = TaigaAPI(host='https://taiga.psuti.ru', auth_type='ldap')
        os.system('cls' if os.name == 'nt' else 'clear')
        try:
            api.auth(
                username=input('Login:\n\r'),
                password=input('Password:\n\r') if not sys.stdin.isatty() else getpass.getpass('Password:\n\r')
            )
            print("\033[A                             \033[A")
            print('\033[92m' + 'authenticated!')

            # Loading animation start
            done = False
            t = threading.Thread(target=animate)
            t.start()

            open('web/.env', 'w').close()
            with open('web/.env', 'w') as f:
                f.write('AUTH_TOKEN=')
                f.write(api.token)

            print('\rGetting info from api...\r')
            tasks = api.tasks.list(project=17)

        except:
            print('incorrect password or login')
            exit()

    # tasks.sort(key=lambda x: x.status, reverse=True)
    print('\rsorting...\r')
    tasks = sorted(tasks, key=lambda x: (x.user_story_extra_info["ref"], x.status_extra_info["name"], x.modified_date))

    user_story = api.user_stories.list(project=17)
    user_story.sort(key=lambda x: x.id)

    if len(config['logs']) == 0:
        print(api)
        print(api.auth)

    new_user_story = {}

    for i in range(len(user_story)):
        new_user_story[vars(user_story[i])['id']] = vars(user_story[i])['status_extra_info']
        new_user_story[vars(user_story[i])['id']]['status'] = vars(user_story[i])['status']

    print('\rinformation received...\r')

    if len(config['fullarr']) == 0:
        for i in range(len(tasks)):
            print(vars(tasks[i]))

    monday = 0

    if len(config['all']) != 0:
        monday = datetime.timestamp(
            now - timedelta(days=now.weekday(), hours=now.hour, minutes=now.minute, seconds=now.second))

    if len(config['data']) > 0:
        monday = datetime.timestamp(datetime.strptime(config['data'], '%d-%m-%Y'))

    generateDocument(new_user_story, tasks, config['task'], config['userstory'], monday)
    done = True
